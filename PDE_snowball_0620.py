from random import seed, uniform
from API_Test_v4 import api
import sys
from snowball_util_V4 import *
from dateutil import rrule
from dateutil import parser
import os
import pickle
from copy import deepcopy
from math import *
from matplotlib import pyplot as plt
import pandas as pd
import numpy as np
import openpyxl as vb
from openpyxl import *
import xlrd
import xlwt
import xlsxwriter
import time
import datetime
np.set_printoptions(suppress=True)
# from datetime import *
current_path = os.getcwd()
# sys.path.append('../Code/')

sheet = pd.read_excel('雪球簿记交易对冲参数_v5.xlsx', header=0,sheet_name='簿记系统【BTC】',usecols="A:AM")
sheet.dropna(axis="index", how='all', inplace=True)

#current_state = sheet.iloc[:, [2]].values.tolist()  # 当前状态
current_state = sheet['当前状态'].values.tolist()  # 当前状态

sheet_v2 = pd.read_excel('product_cal_history.xlsx', header=0 , sheet_name='BTC',)
sheet_v2.dropna(axis="index", how='all', inplace=True)
target_cash_delta = 0
product_name_v2 = sheet_v2.iloc[:, [0]].values.tolist()  # 产品名称
base_date_v2 = sheet_v2['基准日期'].values
expiry_date_v2 = sheet_v2['到期日'].values
current_time_v2 = sheet_v2['现在时间点'].values
volatility_v2 = sheet_v2.iloc[:, [4]].values.tolist()    #波动率
symbol_knock_v2 = sheet_v2.iloc[:, [5]].values.tolist()  # 是否敲入


'''sheet_product_cal_history = pd.read_excel('product_cal_history.xlsx', header=0,sheet_name='BTC',)
sheet_product_cal_history.dropna(axis="index", how='all', inplace=True)
product_name_v2 = sheet_product_cal_history['产品名称'].values.tolist()  # product_cal_history.xlsx 产品名称
volatility_v2 = sheet_product_cal_history['波动率'].values.tolist()     # product_cal_history.xlsx 波动率
symbol_knock_v2 = sheet_product_cal_history['敲入敲出'].values.tolist()  # product_cal_history.xlsx 敲入敲出'''

now_time = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())

#从api导入参数
my_wallet = api.wallet()


for i in range(len(my_wallet)):
    if my_wallet[i]['currency'] == 'USDT':
        coins_USDT = float(my_wallet[i]['available']) + float(my_wallet[i]['frozen'])
    elif my_wallet[i]['currency'] == 'BTC':
        coins_BTC = float(my_wallet[i]['available']) + float(my_wallet[i]['frozen'])

def PDE_BTC(number):

    # 从excel读取参数
    product_name = sheet['产品名称'].values.tolist()  # 产品名称
    product_type = sheet['产品类型'].values.tolist()  # 产品类型
    product_issue_value = sheet['产品发行面值'].values.tolist()  # 产品发行面值
    hook_target = sheet['挂钩标的'].values.tolist()  # 挂钩标的
    base_price = sheet['基准价'].values.tolist()  # 基准价
    base_date = sheet['基准日期'].values  # 基准日期
    expiry_date = sheet['到期日'].values  # 到期日
    operate_days = sheet['运行天数'].values.tolist()  # 运行天数
    knock_in_rate = sheet['敲入比例'].values.tolist()  # 敲入比例
    knock_in_price = sheet['敲入价'].values.tolist()  # 敲入价
    knock_out_rate = sheet['敲出比例'].values.tolist()  # 敲出比例
    knock_out_price = sheet['敲出价'].values.tolist()  # 敲出价
    obs_day = []  # 敲出观察日（穷举数据格式）
    obs_day_tmp = sheet.iloc[:, [13]].values.tolist()
    foreign_coupon = sheet['对外票息'].values.tolist()  # 对外票息
    volatility = sheet['波动率'].values.tolist()  # 波动率
    symbol_knock = sheet['是否敲入'].values.tolist()  # 是否敲入
    current_time = sheet['现在时间点'].values    # 现在时间点
    set_coupon_tmp = sheet.iloc[:, [18]].values.tolist()
    set_coupon = []

    #将敲入的是否转化成布尔值
    for i in range(len(symbol_knock)):
        symbol_knock[i] = symbol_knock[i].replace('是', 'True')
        symbol_knock[i] = symbol_knock[i].replace('否', 'False')
        if symbol_knock[i] == 'True':
            symbol_knock[i] = True
        elif symbol_knock[i] == 'False':
            symbol_knock[i] = False

    #将设定票息转换为浮点数
    for i in range(len(set_coupon_tmp)):
        if set_coupon_tmp[i] is not None:
            set_coupon.append(str(set_coupon_tmp[i][0]).split())
        else:
            continue
    print(set_coupon)
    for i in range(len(set_coupon)):
        for j in range(len(set_coupon[i])):
            if set_coupon[i] is not None:
                set_coupon[i][j] = float(set_coupon[i][j].strip('%')) / 100
                print(type(set_coupon[i][j]))
            else:
                continue

    '''print(set_coupon_tmp[0])
    print(type(set_coupon_tmp[0]))
    print(set_coupon_tmp)
    print(set_coupon_tmp[0][0].split())'''
    print(set_coupon)


    start_date = str(pd.to_datetime(base_date[number]).date())
    end_date = str(pd.to_datetime(expiry_date[number]).date())
    pricing_date = time.strftime('%Y-%m-%d', time.localtime(time.time()))

    today = time.strftime('%Y-%m-%d', time.localtime(time.time()))
    today = datetime.datetime.strptime(today, '%Y-%m-%d').date()
    datestart = datetime.datetime.strptime(start_date, '%Y-%m-%d')
    dateend = datetime.datetime.strptime(end_date, '%Y-%m-%d')

    date_series_v1 = []  # 观察日序列未处理版
    date_obs_number = []  # 序号形式的观察日序列
    Days_sub = []  # 观察日之间间隔
    days_all = [datestart.strftime('%Y-%m-%d')]  # 初始至结束所有日期
    date_series = []  # 处理后的观察日序列

    # 将数据转换为日期形式
    for i in range(len(obs_day_tmp)):
        obs_day.append(obs_day_tmp[i][0].split())

    for i in range(len(obs_day)):
        for j in range(len(obs_day[i])):
            obs_day[i][j] = datetime.datetime.strptime(obs_day[i][j], "%Y/%m/%d").date()

    date_series_v1 = []
    days_observe = obs_day[number]

    for i in days_observe:
        date_series_v1.append(i)

    current_date = str(pd.to_datetime(current_time[number]).date())
    date_series_v1_checking = []
    for t in date_series_v1:
        date_series_v1_checking.append(
            int(rrule.rrule(rrule.DAILY, dtstart=datetime.datetime.strptime(current_date, "%Y-%m-%d").date(),
                            until=t).count()))

    # 找到>0的index
    x = np.array(date_series_v1_checking)
    ind = np.where(x > 0)[0][0]


    # 将已经过期的观察日去除
    for i in date_series_v1:
        if (int(rrule.rrule(rrule.DAILY, dtstart=today, until=i).count())) > 0:
            date_series.append(i)

    # 计算观察日之间间隔
    for i in range(len(date_series) - 1):
        Days_sub.append(rrule.rrule(rrule.DAILY, dtstart=date_series[i], until=date_series[i + 1]).count() - 1)

    # 计算每个观察日与距离初始日期的天数
    for i in range(len(date_series)):
        date_obs_number.append(
            int(rrule.rrule(rrule.DAILY, dtstart=datetime.datetime.strptime(start_date, "%Y-%m-%d").date(),
                            until=date_series[i]).count() - 1))

    while datestart < dateend:  # 生成初始日期至结束日期内所有日期
        datestart += datetime.timedelta(days=1)
        days_all.append(datestart.strftime('%Y-%m-%d'))

    # 产品周期时长
    period_last = rrule.rrule(rrule.DAILY, dtstart=datetime.datetime.strptime(start_date, "%Y-%m-%d").date(),
                              until=datetime.datetime.strptime(end_date, "%Y-%m-%d").date()).count()


    # 发行日到定价日的天数
    days_after_issuance = int(
        rrule.rrule(rrule.DAILY, dtstart=datetime.datetime.strptime(start_date, "%Y-%m-%d").date(),
                    until=datetime.datetime.strptime(pricing_date, "%Y-%m-%d")).count() - 1)
    # 当前定价日距离到期日的天数
    days_to_maturity = int(rrule.rrule(rrule.DAILY, dtstart=datetime.datetime.strptime(pricing_date, "%Y-%m-%d").date(),
                                       until=datetime.datetime.strptime(end_date, "%Y-%m-%d")).count() - 1)

    for i in base_date:
        i = str(pd.to_datetime(i).date())
    for i in expiry_date:
        i = str(pd.to_datetime(i).date())
    for i in current_time:
        i = str(pd.to_datetime(i).date())

    #计算持仓指标
    def compute_value():
        value = 0
        my_wallet = api.wallet()
        for bi in my_wallet:
            symbol = bi['currency']
            amount = float(bi['available']) + float(bi['frozen'])
            if symbol == 'BTC':
                price = float(api.market_kline(symbol='BTC-USDT',
                                               period='1min', size=1)['data'][0]['close'])
            if symbol == 'USDT':
                price = 1

            value += price * amount
        return value

    def BTC_Balance():
        my_wallet = api.wallet()
        for coin in my_wallet:
            if coin['currency'] == 'BTC':
                return float(coin['available']) + float(coin['frozen'])

    def delta_theory(delta):
        return compute_value() * delta

    def delta_real():
        return BTC_Balance()

    def value_need(delta):
        return delta_theory(delta) - delta_real()

    def delta_manual(delta_input):
        return delta_theory(delta_input) - delta_real()

    # %% 这部分确定边界条件，函数可自定义
    # 下边界
    def down_bound_cond(option):
        mult = option._multiplier
        r = option._r  # r: 无风险利率
        strike = option._strike  # strike是敲入边界？
        S_min = option.S_min  # S_min: PDE的价格下限？
        duration = option._duration  # duration: 雪球的duration(天)
        return mult * (S_min / strike) / (1 + r * (duration - np.arange(duration + 1)) / 365)  # 为什么下边界需要贴现？难道不就是strike?

    # 上边界
    def up_bound_cond(option):
        mult = option._multiplier
        ts = option._ts  # ts
        p = option._coupon  # p: coupon rate
        r = option._r  # r: 无风险利率
        T = option._duration / 365  # T: 雪球的duration(年)
        duration = option._duration  # duration: 雪球的duration(天)
        ko_count = option._ko_count  # ko_count: 敲出观察的次数
        obs_dts = option._up_ko_obs_dts  # obs_dts: 敲出观察的日期？
        df_tmp = pd.DataFrame(index=np.arange(duration + 1))  # df_tmp: 声明一个装每天的payoff和obs_dts的DataFrame
        df_tmp.loc[obs_dts, 'payout_f'] = mult * (1 + np.array([p * (obs_dt + ko_count) / 365 for obs_dt in obs_dts]))
        df_tmp.loc[obs_dts, 'obs'] = obs_dts
        df_tmp = df_tmp.fillna(method='bfill')  # 用下一个非缺失的补上
        df_tmp['r_count'] = df_tmp.obs - df_tmp.index  # obs_dts减去index，放到r_count这一列
        df_tmp['payout_f'] = df_tmp['payout_f'] / (1 + r * (df_tmp['r_count'] / 365))
        return df_tmp.payout_f.tolist()

    # 敲入终值条件
    def final_ki_cond(option):
        mult = option._multiplier
        xs = np.exp(option._xs)
        S = option._S0
        T = option._duration / 365
        p = option._coupon
        rp = option._rp
        p = p + rp
        strike = option._strike
        ko_count = option._ko_count
        if type(option._up_ko_barriers) == list:
            u_barrier = option._up_ko_barriers[-1]
        else:
            u_barrier = option._up_ko_barriers
        payout_f = np.clip(xs / strike, -inf, 1) * (xs < u_barrier) + \
                   (1 + p * (T + ko_count / 365)) * (xs >= u_barrier)
        payout_f = payout_f * mult
        return payout_f

    # 敲出终值条件
    def final_noki_cond(option):
        mult = option._multiplier
        xs = np.exp(option._xs)
        S = option._S0
        T = option._duration / 365
        p = option._coupon
        rp = option._rp
        p = (rp + p)
        strike = option._strike
        ko_count = option._ko_count
        d_barrier = option._down_ki_barriers
        payout_f = (1 + p * (T + ko_count / 365)) * (xs >= d_barrier) + np.clip(xs / strike, -inf, 1) * (xs < d_barrier)
        payout_f = payout_f * mult
        return payout_f

    # 敲入观察日条件
    def inside_obs_cond(option, rows, cols):
        mult = option._multiplier
        p = option._coupon
        T = option._duration / 365
        ko_count = option._ko_count
        if type(option._up_ko_barriers) == list:
            u_barrier = option._up_ko_barriers[-1]
        else:
            u_barrier = option._up_ko_barriers

        mesh_tmp = option._mesh.copy()
        ts_tmp = np.array(mesh_tmp.iloc[:, cols].columns.tolist())
        xs_tmp = np.array(mesh_tmp.iloc[rows, :].index.tolist())
        return (mult * (1 + p * (ts_tmp[np.newaxis, :] + ko_count / 365))) * (
                    np.exp(xs_tmp[:, np.newaxis]) >= u_barrier)

    # %% 计算票息
    def cal_coupon(option, down_bound_cond, up_bound_cond, inside_obs_cond, final_ki_cond, final_noki_cond, p=100,
                   n_paths=100000, threshold=0.001, display=False, seed_n=999):
        option_copy = deepcopy(option)
        coupon_uplimit = 1
        coupon_downlimit = 0.01
        threshold = threshold
        count = 0
        while True:
            if count > 100:
                break
            coupon_mid = 0.5 * coupon_uplimit + 0.5 * coupon_downlimit
            option_copy.set_coupon(coupon_mid)
            option_copy.price_value_via_pde(down_bound_cond, up_bound_cond, inside_obs_cond, final_ki_cond,
                                            final_noki_cond)
            p_bar = option_copy._mesh[option_copy._mesh.index <= np.log(100)].iloc[0, 0]
            if display:
                print(count, p_bar, coupon_mid, coupon_downlimit, coupon_uplimit)

            count += 1

            if (p_bar - p) / p < - threshold:
                coupon_downlimit = coupon_mid
            elif (p_bar - p) / p > threshold:
                coupon_uplimit = coupon_mid
            else:
                break

        return coupon_mid, p_bar

    #%% 计算发行日当天估值(票息采用对外售卖的)
    def cal_price_on_issue_day(a,coupon_set,down_bound_cond,up_bound_cond,inside_obs_cond,final_ki_cond,final_noki_cond):
        option_copy = deepcopy(a)
        S0 = option_copy._S0 # p0更新为当天的价格(即不一定为100)
        option_copy._coupon = coupon_set
        option_copy.price_value_via_pde(down_bound_cond,up_bound_cond,inside_obs_cond,final_ki_cond,final_noki_cond)
        option_copy._price_on_issue_day = option_copy._mesh[option_copy._mesh.index>=np.log(S0)].iloc[-1,0]
        return option_copy._price_on_issue_day

    #%% 判断是否计算过
    def product_not_exist(product_name,base_date,expiry_date,current_time,vol,sym_ki):
        flag = 1
        '''products_exist = pd.read_excel('product_cal_history.xlsx', header = None, sheet_name= "BTC")
        product_list_cal_history=products_exist.iloc[:, :].values.tolist()'''

        for i in range(len(product_name_v2)):
            if product_name == int(product_name_v2[i][0]):
                if vol == volatility_v2[i][0]:
                    if sym_ki == symbol_knock_v2[i][0]:
                        if base_date ==base_date_v2[number]:
                            if expiry_date == expiry_date_v2[number]:
                                if current_time == current_time_v2[number]:
                                    flag = 0
                                else:
                                    continue
                            else:
                                continue
                        else:
                            continue
                    else:
                        continue
                else:
                    continue
            else:
                continue
        return flag

    # %% 计算希腊值
    def cal_greeks(a, days_after_issuance,down_bound_cond, up_bound_cond, inside_obs_cond, final_ki_cond,
                   final_noki_cond):
        p0 = a._S0
        und = a._und
        vol = a._vol
        a_vup = deepcopy(a)
        a_vup.set_und(p0, und, vol + 0.01)
        a_vdown = deepcopy(a)
        a_vdown.set_und(p0, und, vol - 0.01)
        a.price_value_via_pde(down_bound_cond, up_bound_cond, inside_obs_cond, final_ki_cond, final_noki_cond)
        a_vup.price_value_via_pde(down_bound_cond, up_bound_cond, inside_obs_cond, final_ki_cond, final_noki_cond)
        a_vdown.price_value_via_pde(down_bound_cond, up_bound_cond, inside_obs_cond, final_ki_cond, final_noki_cond)

        r_change_pct = 0.01
        a_rup = deepcopy(a)
        a_rup._r = a._r + r_change_pct
        a_rdown = deepcopy(a)
        a_rdown._r = a._r - r_change_pct
        a_rup.price_value_via_pde(down_bound_cond, up_bound_cond, inside_obs_cond, final_ki_cond, final_noki_cond)
        a_rdown.price_value_via_pde(down_bound_cond, up_bound_cond, inside_obs_cond, final_ki_cond, final_noki_cond)

        # 求greeks
        f1 = a._mesh.shift(1).copy()
        f2 = a._mesh.copy()
        f3 = a._mesh.shift(-1).copy()
        a.delta_mesh = (f1 - f3) / f2 / (2 * a._dx)
        a.delta_mesh = a.delta_mesh.iloc[1:-1, :]
        a.charm_mesh = (-a.delta_mesh + a.delta_mesh.shift(-1, axis=1)) / (a._dt)
        a.gamma_mesh = (f1 - 2 * f2 + f3) / f2 / f2 / (a._dx ** 2)  # -(f1-f3)/(2*a._dx)/f2/f2
        a.gamma_mesh = a.gamma_mesh.iloc[1:-1, :]
        # theta和vega可能得除以最开始的值
        a.theta_mesh = (-a._mesh + a._mesh.shift(-1, axis=1))
        a.vega_mesh = (a_vup._mesh - a_vdown._mesh) / 0.02 / a._mesh
        a.rho_mesh = (a_rup._mesh - a_rdown._mesh) / 0.02 / a._mesh
        
        delta_mesh = a.delta_mesh
        gamma_mesh = a.gamma_mesh
        theta_mesh = a.theta_mesh
        vega_mesh = a.vega_mesh
        rho_mesh = a.rho_mesh
        
        delta_column = delta_mesh.iloc[:,days_after_issuance]
        gamma_column = gamma_mesh.iloc[:,days_after_issuance]
        theta_column = theta_mesh.iloc[:,days_after_issuance]
        vega_column = vega_mesh.iloc[:,days_after_issuance]
        rho_column = rho_mesh.iloc[:,days_after_issuance]
        
        # 将这些列合并为dataframe，以便写入excel
        dict_delta =  {'delta':delta_column.values}
        df_delta = pd.DataFrame(dict_delta)
        dict_gamma =  {'gamma':gamma_column.values}
        df_gamma = pd.DataFrame(dict_gamma)
        dict_theta =  {'theta':theta_column.values}
        df_theta = pd.DataFrame(dict_theta)
        dict_vega =  {'vega':vega_column.values}
        df_vega = pd.DataFrame(dict_vega)
        dict_rho =  {'rho':rho_column.values}
        df_rho = pd.DataFrame(dict_rho)     

        df_delta.index = delta_mesh.index
        df_gamma.index = gamma_mesh.index
        df_theta.index = theta_mesh.index
        df_vega.index = vega_mesh.index
        df_rho.index = rho_mesh.index

        df_greeks = pd.concat([df_delta,df_gamma,df_theta,df_vega,df_rho],axis=1)
        return df_greeks

    
    def find_greeks(und,greeks_mesh):
        p0 = und
        try:
            '''
            p_plus: 比log(100)稍大一点点的价格
            p_minus: 比log(100)稍小一点点的价格
            d_plus: delta矩阵对应 比log(100)稍大一点点的价格 的delta
            d_minus: delta矩阵对应 比log(100)稍小一点点的价格 的delta
            c_plus: charm矩阵对应 比log(100)稍大一点点的价格 的charm
            c_minus: charm矩阵对应 比log(100)稍小一点点的价格 的charm
            t_plus: theta矩阵对应 比log(100)稍大一点点的价格 的theta
            t_minus: delta矩阵对应 比log(100)稍小一点点的价格 的theta
            v_plus: theta矩阵对应 比log(100)稍大一点点的价格 的vega
            v_minus: delta矩阵对应 比log(100)稍小一点点的价格 的vega
            r_plus = rho矩阵对应 比log(100)稍大一点点的价格 的rho
            r_minus = rho矩阵对应 比log(100)稍小一点点的价格 的rho
            '''
            p_plus = greeks_mesh[greeks_mesh.index >= p0].index[-1]
            p_minus = greeks_mesh[greeks_mesh.index <= p0].index[0]
    
            
            d_plus = greeks_mesh[greeks_mesh.index >= p0].iloc[-1, 1]
            d_minus = greeks_mesh[greeks_mesh.index <= p0].iloc[0, 1]
            g_plus = greeks_mesh[greeks_mesh.index >= p0].iloc[-1, 2]
            g_minus = greeks_mesh[greeks_mesh.index <= p0].iloc[0, 2]
            t_plus = greeks_mesh[greeks_mesh.index >= p0].iloc[-1, 3]
            t_minus = greeks_mesh[greeks_mesh.index <= p0].iloc[0, 3]
            v_plus = greeks_mesh[greeks_mesh.index >= p0].iloc[-1, 4]
            v_minus = greeks_mesh[greeks_mesh.index <= p0].iloc[0, 4]
            r_plus = greeks_mesh[greeks_mesh.index >= p0].iloc[-1, 5]
            r_minus = greeks_mesh[greeks_mesh.index >= p0].iloc[0, 5]
    
            if p_plus == p_minus: # 边缘的直接取值
                delta = d_plus
                gamma = g_plus
                theta = t_plus
                vega = v_plus
                #charm = c_plus
                rho = r_plus
            else: # 内部的插值计算(在excel里完成)
                delta = (p_plus - p0) / (p_plus - p_minus) * d_minus + (p0 - p_minus) / (p_plus - p_minus) * d_plus
                gamma = (p_plus - p0) / (p_plus - p_minus) * g_minus + (p0 - p_minus) / (p_plus - p_minus) * g_plus
                theta = (p_plus - p0) / (p_plus - p_minus) * t_minus + (p0 - p_minus) / (p_plus - p_minus) * t_plus
                vega = (p_plus - p0) / (p_plus - p_minus) * v_minus + (p0 - p_minus) / (p_plus - p_minus) * v_plus
                rho = (p_plus - p0) / (p_plus - p_minus) * r_minus + (p0 - p_minus) / (p_plus - p_minus) * r_plus

        except:
            delta = 0
            gamma = 0
            theta = 0
            vega = 0
            charm = 0
            rho = 0

        return delta, gamma, theta, vega, rho
    
    def find_value(und,today_mesh):
        St = und
        try:
            '''
            p_plus: 比log(100)稍大一点点的价格
            p_minus: 比log(100)稍小一点点的价格
            '''

            p_plus = today_mesh[today_mesh.index >= St].index[-1]
            p_minus = today_mesh[today_mesh.index <= St].index[0]

            price_plus = today_mesh[today_mesh.index >= St].iloc[-1, 0]
            price_minus = today_mesh[today_mesh.index <= St].iloc[0, 0]

            if p_plus == p_minus: # 边缘的直接取值
                now_price = price_plus
            else: # 内部的插值计算(在excel里完成)
                now_price = (p_plus - St) / (p_plus - p_minus) * price_minus + (St - p_minus) / (p_plus - p_minus) * price_plus
        except:
            now_price = 0
        return now_price
    
    # %% 雪球设定
    # obs_dts_df需要先用generate_obs_dts生成或用自定义txt文件
    S0 = 100
    #ub = 105  # 敲出价
    ub = 100 * knock_out_rate[number]
    #db = 65  # 敲入价
    db = 100 * knock_in_rate[number]
    r = 0.015  # 无风险利率
    rp = 0.00  # 固定票息，无论是否敲入敲出
    q = 0.00  # 期限利差或分红
    obs_idx = [i for i in range(len(obs_day))] # 敲出观察日的个数
    ub_dts = [ub for i in range(len(obs_idx))]
    
    # PDE基本信息设定
    gammas = []
    # for und in np.arange(60,110): # 定价时标的价格
    coupon_initial = 0.6442  # 票息初值
    # obs_dts_count = obs_dts_count-days_after_issuance  # 注意obs_dts_count是站在起始日，计算往后的敲出观察日距离起始日有多少天，此处是做了个平移，即计算敲出观察日离现在这个估值日有多久
    obs_dts_count = np.array(date_obs_number)
    
    #ki_flag = False  # 是否已敲入
    ki_flag = symbol_knock[number]
    #vol = 0.60  # 定价用波动率
    vol = volatility[number]
    rate = base_price[number] / 100
    #und = 100   #BTC最新价
    api_data = api.market_kline('BTC-USDT', '1min', 10)

    BTC_latest_price = float(api_data['data'][0]['close'])

    #und = BTC_latest_price[number][0] / rate
    und = BTC_latest_price / rate

    # 上下界条件设定
    a = snowball_option(obs_dts_count[-1], coupon_initial, ko_count=0, ki_flag=ki_flag, r=r, rp=rp, q=q)  # 记录这个雪球的属性
    a.set_up_ko_barrier(ub_dts, obs_dts_count, 'discrete')  # 敲出观察是离散的
    a.set_down_ki_barrier(db, [], 'continuous')  # 敲入观察是连续的
    a.set_und(S0, und, vol)  # 设定最基本的und和vol之类的
    
    # S_max = np.exp((M-int(M*5/6))*dx)*100
    S_min = 10
    M = 1000
    S_max = 500
    a.S_max = S_max
    a.M = M
    a.S_min = S_min
    dx = np.log(100/S_min)/(int(M*5/6))
    a.dx = dx
    a._und = und

    # %% MC部分
    '''
    a.set_mcpaths()
    a.price_value_via_mc()
    print('MC value: %f'%a._value_mc)
    '''

    # %% PDE部分
    if product_not_exist(int(product_name[number]),base_date[number],expiry_date[number],current_time[number],volatility[number],symbol_knock[number]):
        print("该产品之前未计算过，正在执行计算")
        # 发行日需根据卖出的票息计算一个价格(issue price)
        coupon_set = foreign_coupon[number] # 主excel里读
        issue_price = cal_price_on_issue_day(a,coupon_set, down_bound_cond, up_bound_cond, inside_obs_cond, final_ki_cond, final_noki_cond) # 写入主副excel
        # 事实上这个issue_price之后再也不会改变
        # 发行日利用二分法求实际票息(coupon)
        # 采用PDE计算出发行日的coupon，以及对应的价格(由于Linspace未必是100，但是是离100最近的，设为pbar)
        coupon = cal_coupon(a,down_bound_cond,up_bound_cond,inside_obs_cond,final_ki_cond,final_noki_cond)[0]
        pbar = cal_coupon(a,down_bound_cond,up_bound_cond,inside_obs_cond,final_ki_cond,final_noki_cond)[1]
        a._coupon = coupon # a._coupon替换为二分法得到的真实票息
        # 利用真实票息得到真实的估值矩阵
        a.price_value_via_pde(down_bound_cond,up_bound_cond,inside_obs_cond,final_ki_cond,final_noki_cond)
        mesh_copy = a._mesh.copy()
        mesh_copy.index = np.exp(mesh_copy.index)
        mesh_copy.columns = np.linspace(0,period_last,period_last)
        price_column_today = mesh_copy.iloc[:,days_after_issuance]
        price_column_today_dict = {'price':price_column_today.values}
        price_column_today_dict_df = pd.DataFrame(price_column_today_dict)
        price_column_today_dict_df.index = mesh_copy.index
        now_price = mesh_copy[mesh_copy.index>=a._und].iloc[-1,days_after_issuance] # 对应AM列：最新估值
        # 计算Greeks时，采用coupon即内部的
        df_greeks = cal_greeks(a,days_after_issuance,down_bound_cond,up_bound_cond,inside_obs_cond,final_ki_cond,final_noki_cond) # 写入副excel
        df_greeks = df_greeks.iloc[::-1]
        df_greeks.index = mesh_copy.index
        df_output = pd.concat([price_column_today_dict_df,df_greeks],axis=1)

        '''if os.path.exists('output.xlsx'):
            print('1')
        else:
            print('2')
        # 将以上mesh写入到草稿excel中
        wb = vb.load_workbook("output.xlsx")
        writer = pd.ExcelWriter('output.xlsx',engine ='openpyxl')
        writer.book = wb
        book = load_workbook(writer.path)
        writer.book = book
        df_output.to_excel(excel_writer=writer, sheet_name= str(int(product_name[number])))
        writer.save()
        writer.close()'''


        '''wb = vb.load_workbook("output.xlsx")
        ew = pd.ExcelWriter("output.xlsx", engine='openpyxl')
        ew.book = wb
        df_output.to_excel(ew, sheet_name=str(int(product_name[number])))
        ew.save()'''

        wb = vb.load_workbook("output.xlsx")
        ew = pd.ExcelWriter("output.xlsx",engine ='openpyxl')
        ew.book = wb
        #sheets_names = wb.get_sheet_names()
        sheets_names = wb.sheetnames

        if str(int(product_name[number])) in sheets_names:
            ws = wb[str(int(product_name[number]))]
            wb.remove(ws)
            df_output.to_excel(ew, sheet_name= str(int(product_name[number])))
            ew.save()
        else:
            df_output.to_excel(ew, sheet_name= str(int(product_name[number])))
            ew.save()     

        delta, gamma, theta, vega, rho = find_greeks(und, df_output)


        df_history = pd.read_excel('product_cal_history.xlsx',sheet_name='BTC', header=0, index_col=0)

        if product_name[number] in df_history.index:  # 若该产品之前计算过，则覆盖
            print("该产品之前计算过，本次修改参数,正重新计算")
            df_history.loc[product_name[number]] = [base_date[number],expiry_date[number],current_time[number], volatility[number], symbol_knock[number] ]
        else:  # 若该产品没计算过，则加在最后一行
            print("该产品未计算过")
            df_history = pd.concat([df_history, pd.DataFrame(data={'基准日期': base_date[number],
                                                                   '到期日':expiry_date[number],
                                                                   '现在时间点':current_time[number],
                                                                   'vol': volatility[number],
                                                                   'knock': symbol_knock[number]},
                                                             index=[int(product_name[number])],
                                                             columns=df_history.columns)])
        df_history.to_excel("product_cal_history.xlsx", sheet_name='BTC')

        # 输出 期初期权估值 至excel
        wb = vb.load_workbook("雪球簿记交易对冲参数_v5.xlsx")
        ws = wb["簿记系统【BTC】"]
        cell_issue_price = ws.cell(number + 2, 39)
        cell_issue_price.value = issue_price
        # 输出 实际票息 至excel
        cell_actual_coupon = ws.cell(number + 2, 20)
        cell_actual_coupon.value = coupon

        wb.save("雪球簿记交易对冲参数_v5.xlsx")
        '''
        print('\n雪球基本信息: \n 初始价格为{:}, \n 敲出价为{:}, \n 敲入价为{:}, \n 距离到期日还有{:}天, \n 当前标的资产价格为{:}, \n 当前敲入情况(标签)为{:}, \n 波动率为{:}'.format(S0,ub,db,days_to_maturity,und,ki_flag,vol))
        print('\nPDE计算相关参数： \n 价格上限为{:}, \n PDE的列数为{:}, \n PDE的行数(即价格格点数)为{:}'.format(S_max,mesh_copy.shape[1],mesh_copy.shape[0]))
        print('\n[仅用于发行产品时确定隐含票息]票息计算结果: \n 给定估值{:.4}时，coupon为{:.2%}'.format(pbar,coupon))
        print('\n发行第一天根据卖出给客户的票息{:.1%}所计算的估值(此后不再发生变化)为{:.5}'.format(coupon_set,issue_price))
        print('\n希腊字母计算结果: \n delta为{:.4},\n gamma为{:.4},\n theta为{:.4},\n vega为{:.4},\n rho为{:.4}'.format(delta,gamma,theta,vega,rho))
        '''
    else:
        print("该模型已计算过,正在查表")
        # 从草稿excel里读price_mesh, delta_mesh, gamma_mesh, theta_mesh, vega_mesh, rho_mesh)
        # 查表操作
        sheet_name_v1 = str(int(product_name[number]))
        product_mesh = pd.read_excel('output.xlsx', header=0, index_col = 0, sheet_name= sheet_name_v1)
        product_mesh.dropna(axis="index", how='all', inplace=True)

        now_price = find_value(und, product_mesh)
        delta,gamma,theta,vega,rho = find_greeks(und, product_mesh)
        # 以上几个列为希腊字母列series，需放到另一个sheet中
        '''
        print('\n雪球基本信息: \n 初始价格为{:}, \n 敲出价为{:}, \n 敲入价为{:}, \n 距离到期日还有{:}天, \n 当前标的资产价格为{:}, \n 当前敲入情况(标签)为{:}, \n 波动率为{:}'.format(S0,ub,db,days_to_maturity,und,ki_flag,vol))
        print('\nPDE计算相关参数： \n 价格上限为{:}, \n PDE的列数为{:}, \n PDE的行数(即价格格点数)为{:}'.format(S_max,mesh_copy.shape[1],mesh_copy.shape[0]))
        print('\n[计算实时估值，任意一天均可用, coupon固定为在发行日当天使得估值为100的隐含票息]')
        print('\n 给定coupon为{:.5}时，估值计算结果为: {:}'.format(a._coupon,now_price))
        print('\n [计算Greeks时，均采用发行日当天反推所得的隐含票息所得到的PDE矩阵所计算的希腊字母，并取对应价格(对应PDE的行)和对应天(对应PDE的列)]')
        print('\n希腊字母计算结果: \n delta为{:.4},\n gamma为{:.4},\n theta为{:.4},\n vega为{:.4},\n rho为{:.4}'.format(delta,gamma,theta,vega,rho))
        '''
# 对excel写入操作

    days_to_next_obs = int(rrule.rrule(rrule.DAILY, dtstart=today, until=date_series[0]).count() - 1)
    wb = vb.load_workbook("雪球簿记交易对冲参数_v5.xlsx")
    ws = wb["簿记系统【BTC】"]
    # 输出 距离下一观察日天数 至excel
    cell_days_to_next_obs = ws.cell(number + 2, 23)
    cell_days_to_next_obs.value = days_to_next_obs
    # 输出 应计Delta Std 至excel
    cell_delta = ws.cell(number + 2, 25)
    cell_delta.value = delta
    # 输出 实际持仓币数 至excel
    #cell_btc_amount = ws.cell(number + 2, 27)
    #cell_gamma.value = coins_BTC
    # 输出 Gamma Std 至excel
    cell_gamma = ws.cell(number + 2, 32)
    cell_gamma.value = gamma
    # 输出 1%vega 至excel
    cell_vega = ws.cell(number + 2, 35)
    cell_vega.value = vega
    # 输出 theta（1Day) 至excel
    cell_theta = ws.cell(number + 2, 36)
    cell_theta.value = theta
    # 输出 Rho 至excel
    cell_Rho = ws.cell(number + 2, 37)
    cell_Rho.value = rho
    # 输出 最新估值 至excel
    cell_now_price = ws.cell(number + 2, 40)
    cell_now_price.value = now_price

    wb.save("雪球簿记交易对冲参数_v5.xlsx")

for i in range(len(current_state)):
    print("---------------产品", i + 1, "计算结果----------------")
    if (current_state[i] == '已结清'):
        print("该产品已结清，不计算")
        continue
    else:
        print("该产品处于发行中")
        PDE_BTC(i)

wb = vb.load_workbook("雪球簿记交易对冲参数_v5.xlsx")
ws = wb["簿记系统【BTC】"]
# 导入当下时间
cell_now_time = ws.cell(1,47)
cell_now_time.value = now_time
# BTC最新价格
api_data = api.market_kline('BTC-USDT', '1min', 10)
cell_latest_price = ws.cell(2, 47)
cell_latest_price.value = api_data['data'][0]['close']
# 持有USDT数量
cell_USDT_amount = ws.cell(4,47)
cell_USDT_amount.value = coins_USDT
# 持有BTC数量
cell_BTC_amount = ws.cell(5,47)
cell_BTC_amount.value = coins_BTC

wb.save("雪球簿记交易对冲参数_v5.xlsx")


